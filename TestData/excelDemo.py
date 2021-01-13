import openpyxl
book = openpyxl.load_workbook("C:\\Users\\153841\\Desktop\\Python\\SeleniumAutomation\\TestData\\testData.xlsx")
sheet = book.active
cell = sheet.cell(row=1, column=2)
datadict = {}
finaldata = []

for i in range(2, sheet.max_row+1):
    for j in range(2, sheet.max_column+1):
        datadict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
    finaldata.append(datadict)
    datadict = {}




