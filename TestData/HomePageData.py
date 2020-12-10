import openpyxl


class HomePageData:

    def __init__(self,datafile):
        self.datafile = datafile
    #If you want to call the below method via class.method, mark the method as static
    #@staticmethod
    def read_Datafromexcel(self): #self is not required when you declare the method as static
        book = openpyxl.load_workbook(self.datafile)
        sheet = book.active
        cell = sheet.cell(row=1, column=2)
        datadict = {}
        finaldata = []

        for i in range(2, sheet.max_row + 1):
            for j in range(2, sheet.max_column + 1):
                datadict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
            finaldata.append(datadict)
            datadict = {}

        return finaldata